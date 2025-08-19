import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import io

st.title("Asystent Delegacji COQUELLE")

st.info("Wgraj pliki `history.xlsx` oraz `delegacje wzór.xlsx`")

history_file = st.file_uploader("Wgraj plik history.xlsx", type="xlsx")
template_file = st.file_uploader("Wgraj plik delegacje wzór.xlsx", type="xlsx")

def process_file(history_file, template_file):
    # Wczytanie danych history
    history_df = pd.read_excel(history_file, engine="openpyxl")
    history_df = history_df[history_df["Activité du conducteur"] == "Travail"]
    history_df["Date/Heure"] = pd.to_datetime(history_df["Date/Heure"])
    history_df = history_df[(history_df["Date/Heure"].dt.hour >= 7) & (history_df["Date/Heure"].dt.hour <= 18)]
    
    if history_df.empty:
        st.warning("Brak danych w podanym przedziale godzinowym!")
    
    # Podsumowanie dzienne
    daily_summary = {}
    for day, group in history_df.groupby(history_df["Date/Heure"].dt.day):
        earliest = group.loc[group["Date/Heure"].idxmin()]
        latest = group.loc[group["Date/Heure"].idxmax()]
        daily_summary[day] = {
            "earliest_time": earliest["Date/Heure"].strftime("%H:%M"),
            "earliest_country": earliest["Code pays"],
            "latest_time": latest["Date/Heure"].strftime("%H:%M"),
            "latest_country": latest["Code pays"],
        }

    # Wczytanie pliku szablonu
    wb = load_workbook(template_file)
    ws = wb.active

    # Ustawienie miesiąca i roku w J4
    if not history_df.empty:
        first_date = history_df["Date/Heure"].iloc[0]
        ws.cell(row=4, column=10).value = first_date.strftime("%m/%Y")
        last_day_in_data = history_df["Date/Heure"].dt.day.max()
        month = first_date.month
        year = first_date.year
    else:
        today = datetime.today()
        first_date = datetime(today.year, today.month, 1)
        last_day_in_data = 0
        month = today.month
        year = today.year

    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")

    # Liczba dni w miesiącu
    next_month = first_date.replace(day=28) + timedelta(days=4)
    last_day_of_month = (next_month - timedelta(days=next_month.day)).day

    for row in range(8, 8 + last_day_of_month):
        day_cell = ws.cell(row=row, column=1)
        day = row - 7  # bo pierwsza linia z dniem to 8
        weekday = datetime(year, month, day).weekday()

        if weekday >= 5:  # sobota/niedziela
            for col in range(1, 14):
                ws.cell(row=row, column=col).fill = gray_fill
            continue

        if day in daily_summary:
            ws.cell(row=row, column=3).value = daily_summary[day]["earliest_time"]
            ws.cell(row=row, column=4).value = daily_summary[day]["earliest_country"] + " R"
            ws.cell(row=row, column=5).value = daily_summary[day]["latest_time"]
            ws.cell(row=row, column=7).value = daily_summary[day]["latest_country"] + " Z"
        elif day > last_day_in_data:
            ws.cell(row=row, column=2).value = "U"  # brak danych

    # Pobranie wartości z B2 w pliku history
    history_wb = load_workbook(history_file, data_only=True)
    history_ws = history_wb.active
    file_name_from_history = str(history_ws["B2"].value)
    ws["E7"] = file_name_from_history

    # Zapis do BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, file_name_from_history

if st.button("Przetwórz pliki"):
    if history_file and template_file:
        result, filename = process_file(history_file, template_file)
        st.success("Plik przetworzony!")
        st.download_button(
            label="Pobierz wynik",
            data=result,
            file_name=f"{filename}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("Proszę wgrać oba pliki!")
